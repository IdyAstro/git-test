from urllib import request
from django.shortcuts import render , get_object_or_404, redirect # pyright: ignore[reportMissingModuleSource]
from .models import *
from .forms import *
from django.db.models import Q # pyright: ignore[reportMissingModuleSource]
# Create your views here.
from django.contrib.auth.decorators import login_required # pyright: ignore[reportMissingModuleSource]
from .models import *
import csv
from django.http import HttpResponse # pyright: ignore[reportMissingModuleSource]
from django.http import HttpResponse # pyright: ignore[reportMissingModuleSource]
from openpyxl import Workbook
from .models import Demande
from django.db.models import Sum
from datetime import  timedelta
from datetime import datetime
from django.core.paginator import Paginator
from .models import Materiel
from .forms import MaterielForm ,ApproMaterielFormSet,ApprovisionnementForm,DemandeForm
from django.contrib.auth import login , authenticate , logout
from django.contrib import messages 

# Create your views here.
def inscription(reqest):
      if reqest.method == "POST":
        forms = CustomUserCreationForm(reqest.POST)
        if forms.is_valid():
           forms.save()
           return redirect('connexion')
      else:
        forms = CustomUserCreationForm()
      return render(reqest,'inscription.html',{'forms':forms})



def connexion(request):
    if request.method == "POST":

        username = request.POST.get("username")
        password = request.POST.get("password")
        user = authenticate(request, username=username, password=password)
        # vérification des identifiants ici
        request.session['user_id'] = user.id  # stocke l'id dans la session
        request.session['is_authenticated'] = True

        if user is not None:
            login(request, user)
            return redirect("dashboard")
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect ❌")

    # Ici, GET ou POST échoué → juste afficher le template
    return render(request, "login.html")
from django.contrib.auth import logout
from django.shortcuts import redirect

def deconnexion(request):
    logout(request)
    return redirect("connexion")  # Redirige vers la page de connexion après déconnexion



@login_required(login_url='connexion')
def home(request):
    # Stats simples
   #stockTotal = int(Stock.objects.aaggregate(sum('quantite')))

    # Evolution 7 derniers jours
    
   return render(request, "home.html")


from django.shortcuts import render, redirect, get_object_or_404
from .models import Categorie
from django.contrib import messages
#=========================CATEGORIE================================================================
# ---- Liste ----
@login_required(login_url='connexion')
def liste_categories(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    categories = Categorie.objects.all().order_by('id')
    return render(request, 'categories/liste.html', {'categories': categories})

# ---- Ajouter ----
@login_required(login_url='connexion')
def ajouter_categorie(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    if request.method == 'POST':
        label = request.POST.get('label')
        if Categorie.objects.filter(label__iexact=label).exists():
            messages.error(request, "Cette catégorie existe déjà.")
        else:
            Categorie.objects.create(label=label)
            messages.success(request, "Catégorie ajoutée avec succès.")
            return redirect('liste_categories')
    return render(request, 'categories/form.html')

# ---- Modifier ----
@login_required(login_url='connexion')
def modifier_categorie(request, id):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    categorie = get_object_or_404(Categorie, id=id)
    if request.method == 'POST':
        label = request.POST.get('label')
        if label:
            categorie.label = label
            categorie.save()
            messages.success(request, "Catégorie modifiée avec succès !")
            return redirect('liste_categories')
    return render(request, 'categories/form.html', {'categorie': categorie})

# ---- Supprimer ----
@login_required(login_url='connexion')
def supprimer_categorie(request, id):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    categorie = get_object_or_404(Categorie, id=id)
    categorie.delete()
    messages.success(request, "Catégorie supprimée !")
    return redirect('liste_categories')
#=============================MATERIEL=============================================================





# ---- LISTE avec filtres et pagination ----
@login_required(login_url='connexion')
def materiel_list(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    query = request.GET.get("q")
    categorie = request.GET.get("categorie")

    materiels = Materiel.objects.all()

    if query:
        materiels = materiels.filter(fabricant__icontains=query) | materiels.filter(description__icontains=query) | materiels.filter(modal__icontains=query) 

    if categorie:
        materiels = materiels.filter(categorie__id=categorie)

    paginator = Paginator(materiels, 5)  # 5 matériels par page
    page = request.GET.get("page")
    materiels_page = paginator.get_page(page)

    return render(request, "materiel/liste.html", {"materiels": materiels_page})

# ---- CREATE ----
@login_required(login_url='connexion')
def materiel_create(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    if request.method == "POST":
        form = MaterielForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("materiel_list")
    else:
        form = MaterielForm()
    return render(request, "materiel/form.html", {"form": form})

# ---- UPDATE ----
@login_required(login_url='connexion')
def materiel_update(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    materiel = get_object_or_404(Materiel, pk=pk)
    if request.method == "POST":
        form = MaterielForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
            return redirect("materiel_list")
    else:
        form = MaterielForm(instance=materiel)
    return render(request, "materiel/modif.html", {"form": form})

# ---- DELETE ----
@login_required(login_url='connexion')
def materiel_delete(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    materiel = get_object_or_404(Materiel, pk=pk)
    if request.method == "POST":
        materiel.delete()
        return redirect("materiel_list")
    return render(request, "materiel/confirme_suppression.html", {"materiel": materiel})



# views.py


#====================================APPROVISIONNEMENT=======================================================================
@login_required(login_url='connexion')
def liste_appros(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    
    appros = Approvisionnement.objects.all().order_by('-date_appro')
    paginator = Paginator(appros, 10)
    page = request.GET.get("page")
    appros_page = paginator.get_page(page)
    return render(request, "appro/liste.html", {"appros": appros_page})

@login_required(login_url='connexion')
def historique_appros(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    appros = Approvisionnement.objects.all().order_by('-date_appro')
    paginator = Paginator(appros, 10)
    page = request.GET.get("page")
    appros_page = paginator.get_page(page)
    return render(request, "appro/historique.html", {"appros": appros_page})


@login_required(login_url='connexion')
def creer_appro(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    if request.method == "POST":
        for form in formset:
         form.initial['approvisionnement'] = getattr(form.instance.approvisionnement, 'id', '')

        appro_form = ApprovisionnementForm(request.POST)
        formset = ApproMaterielFormSet(request.POST)

        if appro_form.is_valid() and formset.is_valid():
            appro = appro_form.save()
            formset.instance = appro
            formset.save()
            return redirect("liste_appros")
    else:
        appro_form = ApprovisionnementForm()
        formset = ApproMaterielFormSet()

    return render(request, "appro/form.html", {
        "appro_form": appro_form,
        "formset": formset
    })

@login_required(login_url='connexion')
def modifier_appro(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    appro = get_object_or_404(Approvisionnement, pk=pk)

    if request.method == "POST":
        appro_form = ApprovisionnementForm(request.POST, instance=appro)
        formset = ApproMaterielFormSet(request.POST, instance=appro)

        if appro_form.is_valid() and formset.is_valid():
            appro_form.save()
            formset.save()
            return redirect("liste_appros")
    else:
        appro_form = ApprovisionnementForm(instance=appro)
        formset = ApproMaterielFormSet(instance=appro)

    return render(request, "appro/form.html", {
        "appro_form": appro_form,
        "formset": formset,
        "modifier": True  # utile pour différencier création et modification dans le template
    })

@login_required(login_url='connexion')
def detail_appro(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    appro = get_object_or_404(Appro, pk=pk)
    return render(request, 'appro/detail_appro.html', {'appro': appro})

#===================================================DEMANDE==========================================================
# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Demande
from .forms import DemandeForm

# ---- CREATE ----
@login_required(login_url='connexion')
def creer_demande(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    if request.method == "POST":
        form = DemandeForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect("listes_exhaussive")
    else:
        form = DemandeForm()
    return render(request, "demandes/creer.html", {"form": form})

# ---- READ (Liste) ----
@login_required(login_url='connexion')


@login_required(login_url='connexion')



@login_required(login_url='connexion')
def export_demandes_xls(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')

    # Récupération des filtres GET
    direction = request.GET.get('direction')
    departement = request.GET.get('departement')
    statut = request.GET.get('statut')
    q = request.GET.get('q')

    # Filtrer les demandes
    demandes = Demande.objects.all()
    if direction:
        demandes = demandes.filter(direction=direction)
    if departement:
        demandes = demandes.filter(departement=departement)
    if statut:
        demandes = demandes.filter(statut=statut)
    if q:
        demandes = demandes.filter(Q(prenom__icontains=q) | Q(nom__icontains=q))

    # Créer un workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Demandes"

    # Écrire l’en-tête
    headers = ['Prenom', 'Nom', 'Categorie', 'Description', 'Direction', 'Departement',
               'Service', 'Email', 'Quantite', 'Motif', 'Statut', 'Date']
    ws.append(headers)

    # Ajouter les lignes
    for d in demandes:
        row = [
            d.prenom,
            d.nom,
            str(d.materiel.categorie) if d.materiel else "",
            str(d.materiel.description) if d.materiel else "",
            str(d.direction.nom) if d.direction else "",
            str(d.departement.nom) if d.departement else "",
            str(d.service.nom) if d.service else "",
            d.email,
            d.quantite,
            d.motif,
            d.get_statut_display(),  # Affiche la version lisible du statut
            d.date_demande.strftime("%d/%m/%Y %H:%M") if d.date_demande else "",
        ]
        ws.append(row)

    # Préparer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=demandes.xlsx'
    wb.save(response)
    return response
@login_required(login_url='connexion')
def liste_demandes(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    query = request.GET.get("q")
    demandes = Demande.objects.all().order_by("-date_demande")[:4]
    if query:
        demandes = demandes.filter(prenom__icontains=query) | demandes.filter(nom__icontains=query)
    
    
    paginator = Paginator(demandes, 4)  
    page = request.GET.get("page")
    demandes_page = paginator.get_page(page)
    return render(request, "demandes/liste.html", {"demandes": demandes_page})

# ---- READ (Détail) ----
@login_required(login_url='connexion')
def detail_demande(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    demande = get_object_or_404(Demande, pk=pk)
    return render(request, "demandes/detail.html", {"demande": demande})

# ---- UPDATE ----
@login_required(login_url='connexion')
def modifier_demande(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    demande = get_object_or_404(Demande, pk=pk)
    if request.method == "POST":
        form = DemandeForm(request.POST, instance=demande)
        if form.is_valid():
            form.save()
            return redirect("listes_exhaussive")
    else:
        form = DemandeForm(instance=demande)
    return render(request, "demandes/modifier.html", {"form": form})

# ---- DELETE ----
@login_required(login_url='connexion')
def supprimer_demande(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    demande = get_object_or_404(Demande, pk=pk)
    if request.method == "POST":
        demande.delete()
        return redirect("liste_demandes")
    return render(request, "demandes/supprimer.html", {"demande": demande})

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator

from .models import Demande, Direction, Departement, Service

@login_required(login_url='connexion')
def listes_exhaussive(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')

    query = request.GET.get("q")
    demandes = Demande.objects.all().order_by("-date_demande")
    if query:
        demandes = demandes.filter(Q(prenom__icontains=query) | Q(nom__icontains=query))

    # Filtres dynamiques
    current_filters = {}
    filter_fields = ['direction', 'departement', 'service', 'statut']
    for field in filter_fields:
        value = request.GET.get(field, "")
        if value:
            kwargs = {f"{field}": value}
            demandes = demandes.filter(**kwargs)
        current_filters[field] = value

    paginator = Paginator(demandes, 4)
    page = request.GET.get("page")
    demandes_page = paginator.get_page(page)

    # Champs pour filtres
    radio_fields = [Demande._meta.get_field(f) for f in filter_fields]

    # Préparer dictionnaires pour afficher les noms dans le template
    directions = {str(d.id): d.nom for d in Direction.objects.all()}
    departements = {str(d.id): d.nom for d in Departement.objects.all()}
    services = {str(s.id): s.nom for s in Service.objects.all()}

    # Statistiques
    total_demandes = Demande.objects.count()
    demandes_attente = Demande.objects.filter(statut='en_attente').count()
    demandes_acceptee = Demande.objects.filter(statut='acceptee').count()
    demandes_refusee = Demande.objects.filter(statut='refusee').count()
    chart_labels = ['En attente', 'Validé', 'Refusée']
    chart_data = [
        demandes_attente,
        demandes_acceptee,
        total_demandes - demandes_attente - demandes_acceptee
    ]

    return render(request, "demandes/listes_exhaussive.html", {
        "demandes": demandes_page,
        "radio_fields": radio_fields,
        "current_filters": current_filters,
        "directions": directions,
        "departements": departements,
        "services": services,
        "total_demandes": total_demandes,
        "demandes_attente": demandes_attente,
        "demandes_acceptee": demandes_acceptee,
        "demandes_refusee": demandes_refusee,
        "chart_labels": chart_labels,
        "chart_data": chart_data,
        "request": request,
    })

    
from django.shortcuts import render, redirect
from django.forms import modelformset_factory
from django.contrib import messages
from .models import Attribution, AttribuMateriel
from .forms import AttributionForm, AttribuMaterielForm

from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Attribution

from django.shortcuts import render, redirect
from django.core.paginator import Paginator
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import Attribution
from django.db.models import Q
from django.http import HttpResponse
import csv
from openpyxl import Workbook
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime

@login_required(login_url='connexion')
def liste_attributions(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    
    qs = Attribution.objects.all()

    # Recherche texte
    q = request.GET.get("q", "")
    if q:
        qs = qs.filter(
            Q(prenom__icontains=q) |
            Q(nom__icontains=q) |
            Q(ref__icontains=q)
        )

    # Filtres dynamiques
    current_filters = {}
    filter_fields = ['direction', 'departement', 'service', 'etat']
    for field in filter_fields:
        value = request.GET.get(field, "")
        if value:
            kwargs = {f"{field}": value}
            qs = qs.filter(**kwargs)
        current_filters[field] = value

    # Filtre par année
    annee = request.GET.get('annee')
    if annee:
        qs = qs.filter(date_attri__year=annee)
    current_filters['annee'] = annee

    # Pagination
    paginator = Paginator(qs.order_by("-date_attri"), 5)
    page_number = request.GET.get("page")
    attributions_page = paginator.get_page(page_number)

    # Préparer radio_fields pour template
    radio_fields = [Attribution._meta.get_field(f) for f in filter_fields]

    return render(request, "attributions/liste.html", {
        "attributions": attributions_page,
        "radio_fields": radio_fields,
        "current_filters": current_filters,
        "request": request,
    })
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.core.paginator import Paginator
from django.db.models import Q, Prefetch
from django.http import HttpResponse
import openpyxl
from .models import Attribution, AttribuMateriel

@login_required(login_url='connexion')
def rapport_attributions(request):
    qs = Attribution.objects.prefetch_related(
        Prefetch("attribumateriel_set", queryset=AttribuMateriel.objects.select_related("materiel"))
    )

    # Recherche texte
    q = request.GET.get("q", "")
    if q:
        qs = qs.filter(
            Q(prenom__icontains=q) |
            Q(nom__icontains=q) |
            Q(ref__icontains=q)
        )

    # Filtres dynamiques
    current_filters = {}
    filter_fields = ['direction', 'departement', 'service', 'etat']
    for field in filter_fields:
        value = request.GET.get(field, "")
        if value:
            kwargs = {f"{field}": value}
            qs = qs.filter(**kwargs)
        current_filters[field] = value

    # Filtre par année
    annee = request.GET.get('annee')
    if annee:
        qs = qs.filter(date_attri__year=annee)
    current_filters['annee'] = annee

    # Export Excel
    if request.GET.get("export") == "excel":
        return export_attributions_excel(qs)

    # Pagination
    paginator = Paginator(qs.order_by("-date_attri"), 10)
    page_number = request.GET.get("page")
    attributions_page = paginator.get_page(page_number)

    # Radio fields pour le template
    radio_fields = [Attribution._meta.get_field(f) for f in filter_fields]

    return render(request, "attributions/rapport.html", {
        "attributions": attributions_page,
        "radio_fields": radio_fields,
        "current_filters": current_filters,
        "request": request,
    })


def export_attributions_excel(queryset):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attributions"

    # Entêtes
    ws.append([
        "Date", "Direction", "Département", "Service", "Utilisateur", 
        "Technicien", "Matériel", "Quantité", "État"
    ])

    for attri in queryset:
        for ligne in attri.attribumateriel_set.all():
            ws.append([
                attri.date_attri.strftime("%Y-%m-%d"),
                str(attri.direction),
                str(attri.departement),
                str(attri.service),
                f"{attri.prenom} {attri.nom}",
                str(attri.utilisateur),
                str(ligne.materiel),
                ligne.quantite,
                attri.get_etat_display()
            ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=rapport_attributions.xlsx'
    wb.save(response)
    return response

from django.http import HttpResponse
from openpyxl import Workbook
import csv
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from .models import Attribution
from django.contrib.auth.decorators import login_required
from django.contrib import messages

@login_required(login_url='connexion')
def export_attributions(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')

    # Récupérer le format demandé : xlsx, csv, pdf
    export_format = request.GET.get('format', 'xlsx')

    # Récupérer les filtres GET
    direction = request.GET.get('direction')
    departement = request.GET.get('departement')
    service = request.GET.get('service')
    etat = request.GET.get('etat')
    utilisateur = request.GET.get('utilisateur')
    q = request.GET.get('q')

    # Filtrer les attributions
    qs = Attribution.objects.all()
    if direction:
        qs = qs.filter(direction_id=direction)
    if departement:
        qs = qs.filter(departement_id=departement)
    if service:
        qs = qs.filter(service_id=service)
    if etat:
        qs = qs.filter(etat=etat)
    if utilisateur:
        qs = qs.filter(utilisateur_id=utilisateur)
    if q:
        qs = qs.filter(ref__icontains=q) | qs.filter(prenom__icontains=q) | qs.filter(nom__icontains=q)

    # Liste des colonnes
    headers = ['Date', 'Prénom', 'Nom', 'Référence', 'Direction', 'Département', 'Service', 'État', 'Utilisateur', 'Email', 'Localité', 'Série']

    if export_format == 'xlsx':
        # Création XLSX
        wb = Workbook()
        ws = wb.active
        ws.title = "Attributions"
        ws.append(headers)

        for attr in qs:
            ws.append([
                attr.date_attri.strftime("%d/%m/%Y %H:%M") if attr.date_attri else '',
                attr.prenom,
                attr.nom,
                attr.ref,
                str(attr.direction),
                str(attr.departement),
                str(attr.service),
                attr.etat,
                attr.utilisateur.nom if attr.utilisateur else '',
                attr.email,
                attr.locality,
                attr.serie
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=attributions.xlsx'
        wb.save(response)
        return response

    elif export_format == 'csv':
        # Création CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=attributions.csv'
        writer = csv.writer(response)
        writer.writerow(headers)

        for attr in qs:
            writer.writerow([
                attr.date_attri.strftime("%d/%m/%Y %H:%M") if attr.date_attri else '',
                attr.prenom,
                attr.nom,
                attr.ref,
                attr.direction,
                attr.departement,
                attr.service,
                attr.etat,
                attr.utilisateur.nom if attr.utilisateur else '',
                attr.email,
                attr.locality,
                attr.serie
            ])
        return response

    elif export_format == 'pdf':
        # Création PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        x_margin = 30
        y = height - 50

        # Titre
        p.setFont("Helvetica-Bold", 14)
        p.drawString(x_margin, y, "Liste des deploiements")
        y -= 30

        # En-têtes
        p.setFont("Helvetica-Bold", 10)
        for i, header in enumerate(headers):
            p.drawString(x_margin + i*100, y, header)
        y -= 20

        # Données
        p.setFont("Helvetica", 9)
        for attr in qs:
            row = [
                attr.date_attri.strftime("%d/%m/%Y %H:%M") if attr.date_attri else '',
                attr.prenom,
                attr.nom,
                attr.ref,
                attr.direction,
                attr.departement,
                attr.service,
                attr.etat,
                attr.utilisateur.nom if attr.utilisateur else '',
                attr.email,
                attr.locality,
                attr.serie
            ]
            for i, item in enumerate(row):
                p.drawString(x_margin + i*100, y, str(item))
            y -= 15
            if y < 50:  # nouvelle page
                p.showPage()
                y = height - 50

        p.save()
        pdf = buffer.getvalue()
        buffer.close()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=attributions.pdf'
        return response

    else:
        messages.error(request, "Format d'export non supporté.")
        return redirect('liste_attributions')


@login_required(login_url='connexion')
def attribution_detail(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    attribution = get_object_or_404(Attribution, pk=pk)
    return render(request, "attributions/detail.html", {
        "attribution": attribution,
    })

@login_required(login_url='connexion')
def nouvelle_attribution(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    AttribuFormSet = modelformset_factory(AttribuMateriel, form=AttribuMaterielForm, extra=1, can_delete=True)

    if request.method == "POST":
        attribution_form = AttributionForm(request.POST)
        formset = AttribuFormSet(request.POST, queryset=AttribuMateriel.objects.none())

        if attribution_form.is_valid() and formset.is_valid():
            attribution = attribution_form.save()

            for form in formset:
                if form.cleaned_data and not form.cleaned_data.get("DELETE", False):
                    attribu_materiel = form.save(commit=False)
                    attribu_materiel.attri = attribution
                    try:
                        attribu_materiel.save()
                    except ValueError as e:
                        messages.error(request, str(e))
                        attribution.delete()
                        return redirect("nouvelle_attribution")

            messages.success(request, "✅ Attribution enregistrée avec succès.")
            return redirect("liste_attributions")

    else:
        attribution_form = AttributionForm()
        formset = AttribuFormSet(queryset=AttribuMateriel.objects.none())

    return render(request, "attributions/form.html", {
        "attribution_form": attribution_form,
        "formset": formset,
    })

# Export CSV
@login_required(login_url='connexion')
def exportation_csv(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    attributions = Attribution.objects.all()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attributions.csv"'

    writer = csv.writer(response)
    writer.writerow(['Ref', 'Nom', 'Prenom', 'Email', 'Direction', 'Departement', 'Service', 'Locality', 'Etat', 'Date'])
    for a in attributions:
        writer.writerow([a.ref, a.nom, a.prenom, a.email, a.direction, a.departement, a.service, a.locality, a.etat, a.date_attri])
    return response

@login_required(login_url='connexion')
def edit_attribution(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    attribution = get_object_or_404(Attribution, pk=pk)
    if request.method == "POST":
        form = AttributionForm(request.POST, instance=attribution)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Attribution mise à jour avec succès")
            return redirect('liste_attributions')
    else:
        form = AttributionForm(instance=attribution)
    return render(request, "attributions/edit.html", {"form": form})
from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages
from .models import Attribution
@login_required(login_url='connexion')
def delete_attribution(request, pk):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    attribution = get_object_or_404(Attribution, pk=pk)
    
    if request.method == "POST":
        attribution.delete()
        messages.success(request, f"🗑 Attribution #{pk} supprimée avec succès")
    
    return redirect('liste_attributions')


import json
from django.shortcuts import render
from django.db.models import Sum, Count
from .models import Stock, Demande, ApproMateriel, AttribuMateriel, Materiel
@login_required(login_url='connexion')
def dashboard(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    
    total_materiels = Materiel.objects.count()
    
    total_stock = Stock.objects.aggregate(total=Sum("quantite"))["total"] or 0
    total_appros = ApproMateriel.objects.aggregate(total=Sum("quantite"))["total"] or 0
    total_attris = AttribuMateriel.objects.aggregate(total=Sum("quantite"))["total"] or 0

    demandes_stats = Demande.objects.values("statut").annotate(nb=Count("id"))

    top_appros = (
        ApproMateriel.objects.values("materiel__description")
        .annotate(total=Sum("quantite"))
        .order_by("-total")[:5]
    )
    top_attris = (
        AttribuMateriel.objects.values("materiel__description")
        .annotate(total=Sum("quantite"))
        .order_by("-total")[:5]
    )

    stock_par_categorie = (
        Stock.objects.values("materiel__categorie__label")
        .annotate(total=Sum("quantite"))
    )

    stock_critique = Stock.objects.filter(quantite__lt=5)
    stock_par_materiel = Stock.objects.filter(quantite__gt=0)

    # --- Préparer les listes JSON pour Chart.js ---
    chart_stock_labels = [s["materiel__categorie__label"] for s in stock_par_categorie]
    chart_stock_data = [s["total"] for s in stock_par_categorie]

    chart_demande_labels = [d["statut"] for d in demandes_stats]
    chart_demande_data = [d["nb"] for d in demandes_stats]

    context = {
        "total_stock": total_stock,
        "total_materiels": total_materiels,
        "total_appros": total_appros,
        "total_attris": total_attris,
        "top_appros": list(top_appros),
        "top_attris": list(top_attris),
        "stock_critique": stock_critique,
        "stock_par_materiel": stock_par_materiel,

        # JSON safe
        "chart_stock_labels": json.dumps(chart_stock_labels),
        "chart_stock_data": json.dumps(chart_stock_data),
        "chart_demande_labels": json.dumps(chart_demande_labels),
        "chart_demande_data": json.dumps(chart_demande_data),
    }
    return render(request, "dashboard.html", context)


@login_required(login_url='connexion')
def nouvel_approvisionnement(request):
    # Vérification supplémentaire session
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('connexion')

    # Traitement POST
    if request.method == "POST":
        utilisateur_id = request.POST.get("utilisateur")
        utilisateur = Utilisateur.objects.get(id=utilisateur_id)

        # Création de l'approvisionnement
        appro = Approvisionnement.objects.create(utilisateur=utilisateur)

        # Boucle sur tous les matériels
        for materiel in Materiel.objects.all():
            # Vérifie si la case est cochée
            if request.POST.get(f"check_{materiel.id}"):
                quantite = int(request.POST.get(f"quantite_{materiel.id}", 0) or 0)
                if quantite > 0:
                    ApproMateriel.objects.create(
                        appro=appro,
                        materiel=materiel,
                        quantite=quantite
                    )

        messages.success(request, "Approvisionnement enregistré avec succès ✅")
        return redirect("liste_appros")

    # GET : récupération utilisateurs et matériels
    utilisateurs = Utilisateur.objects.all()
    materiels_list = Materiel.objects.all()  # ✅ QuerySet pour pagination

    # Pagination sur les matériels (10 par page)
    paginator = Paginator(materiels_list, 3)
    page_number = request.GET.get("page")
    materiels_pages = paginator.get_page(page_number)

    return render(request, "appro/nouvel_appro.html", {
        "utilisateurs": utilisateurs,
        "materiels": materiels_pages
    })

from django.shortcuts import render, redirect, get_object_or_404
from .models import Attribution, AttribuMateriel
from .forms import AttributionForm, AttribuMaterielForm
from django.db.models import Q
from django.contrib import messages

from django.http import HttpResponse
import csv

def filtre_attributions(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    attributions = Attribution.objects.all()
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    etat = request.GET.get('etat')
    service = request.GET.get('service')
    locality = request.GET.get('locality')

    if start_date and end_date:
        attributions = attributions.filter(date_attri__date__range=[start_date, end_date])
    if etat:
        attributions = attributions.filter(etat=etat)
    if service:
        attributions = attributions.filter(service=service)
    if locality:
        attributions = attributions.filter(locality=locality)

    return render(request, 'attributions/liste.html', {'attributions': attributions})




@login_required(login_url='connexion')
def rapport_attributions():
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    rapport = (
        Categorie.objects
        .annotate(
            total_appro=Sum("materiel__appromateriel__quantite"),
            total_attri=Sum("materiel__attribumateriel__quantite"),
            stock=Sum("materiel__stock__quantite"),
        )
        .values("label", "total_appro", "total_attri", "stock")
    )
    return rapport

@login_required(login_url='connexion')
def rapport_par_direction():
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    rapport = (
        Attribution.objects
        .values("direction")
        .annotate(total_attribue=Sum("attribumateriel__quantite"))
        .order_by("-total_attribue")
    )
    return rapport


from django.shortcuts import render
from django.db.models import Sum
@login_required(login_url='connexion')
def rapport_global(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    categories = (
        Categorie.objects
        .annotate(
            total_appro=Sum("materiel__appromateriel__quantite"),
            total_attri=Sum("materiel__attribumateriel__quantite"),
            stock=Sum("materiel__stock__quantite"),
        )
    )

    directions = (
        Attribution.objects
        .values("direction")
        .annotate(total_attribue=Sum("attribumateriel__quantite"))
    )

    return render(request, "rapport.html", {
        "categories": categories,
        "directions": directions,
    })

@login_required(login_url='connexion')
def export_rapport_excel(request):
    # Création du fichier Excel
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Attributions"

    # === Section 1 : Synthèse par catégorie ===
    ws.append(["Rapport global des matériels"])
    ws.append(["Catégorie", "Total Approvisionné", "Total Attribué", "Stock Restant"])

    categories = (
        Categorie.objects
        .annotate(
            total_appro=Sum("materiel__appromateriel__quantite"),
            total_attri=Sum("materiel__attribumateriel__quantite"),
            stock=Sum("materiel__stock__quantite"),
        )
    )

    for c in categories:
        ws.append([
            c.label,
            c.total_appro or 0,
            c.total_attri or 0,
            c.stock or 0
        ])

    ws.append([])  # ligne vide

    # === Section 2 : Répartition par direction ===
    ws.append(["Répartition des attributions par Direction"])
    ws.append(["Direction", "Quantité Attribuée"])

    directions = (
        Attribution.objects
        .values("direction")
        .annotate(total_attribue=Sum("attribumateriel__quantite"))
    )

    for d in directions:
        ws.append([d["direction"], d["total_attribue"] or 0])

    ws.append([])

    # === Section 3 : Détail des attributions ===
    ws.append(["Détail des attributions"])
    ws.append([
        "Date Attribution", "Réf", "Employé", "Email",
        "Direction", "Département", "Service",
        "Matériel", "Quantité", "État"
    ])

    attributions = Attribution.objects.all().prefetch_related("attribumateriel_set__materiel")
    for attri in attributions:
        for mat in attri.attribumateriel_set.all():
            ws.append([
                attri.date_attri.strftime("%Y-%m-%d %H:%M"),
                attri.ref,
                f"{attri.prenom} {attri.nom}",
                attri.email,
                attri.direction,
                attri.departement,
                attri.service,
                str(mat.materiel),
                mat.quantite,
                attri.etat,
            ])

    # === Retourner en réponse HTTP ===
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="rapport_attributions.xlsx"'
    wb.save(response)
    return response

from django.shortcuts import render
from django.db.models import Sum
from django.http import HttpResponse
import openpyxl
from .models import Categorie, Attribution

from django.utils.timezone import now
from datetime import timedelta
@login_required(login_url='connexion')
def rapport_complet(request):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    # === Section 1 : Synthèse par catégorie ===
    stock_par_categorie = (
        Stock.objects.values("materiel__categorie__label")
        .annotate(total=Sum("quantite"))
    )
    categories = (
        Categorie.objects
        .annotate(
            total_appro=Sum("materiel__appromateriel__quantite"),
            total_attri=Sum("materiel__attribumateriel__quantite"),
            stock=Sum("materiel__stock__quantite"),
        )
    )

    # === Section 2 : Répartition par direction ===
    directions = (
        Attribution.objects
        .values("direction")
        .annotate(total_attribue=Sum("attribumateriel__quantite"))
        .order_by("-total_attribue")
    )

    # === Section 3 : Détail des attributions ===
    attributions = Attribution.objects.all().prefetch_related("attribumateriel_set__materiel")

    # === Section 4 : Matériels anciens à remplacer ===
    trois_ans = now().date() - timedelta(days=3*365)
    materiels_a_remplacer = Attribution.objects.filter(date_attri__lte=trois_ans)

    # Vérifier si export demandé
    export = request.GET.get("export")
    if export == "excel":
        return export_excel(categories, directions, attributions, materiels_a_remplacer)

    return render(request, "attributions/rapport_complet.html", {
        "categories": categories,
        "directions": directions,
        "attributions": attributions,
        "stock_par_categorie":stock_par_categorie,
        "materiels_a_remplacer": materiels_a_remplacer,
    })

# ====== Fonction export Excel ======
@login_required(login_url='connexion')
def export_excel(categories, directions, attributions):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rapport Attributions"

    # Synthèse par catégorie
    ws.append(["Synthèse par Catégorie"])
    ws.append(["Catégorie", "Approvisionné", "Attribué", "Stock"])
    for c in categories:
        ws.append([c.label, c.total_appro or 0, c.total_attri or 0, c.stock or 0])
    ws.append([])

    # Répartition par direction
    ws.append(["Répartition par Direction"])
    ws.append(["Direction", "Quantité Attribuée"])
    for d in directions:
        ws.append([d["direction"], d["total_attribue"] or 0])
    ws.append([])

    # Détail des attributions
    ws.append(["Détails des Attributions"])
    ws.append([
        "Date", "Réf", "Employé", "Email", "Direction",
        "Département", "Service", "Matériel", "Quantité", "État"
    ])
    for attri in attributions:
        for mat in attri.attribumateriel_set.all():
            ws.append([
                attri.date_attri.strftime("%Y-%m-%d %H:%M"),
                attri.ref,
                f"{attri.prenom} {attri.nom}",
                attri.email,
                attri.direction,
                attri.departement,
                attri.service,
                str(mat.materiel),
                mat.quantite,
                attri.etat,
            ])

    # Retour HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="rapport_attributions.xlsx"'
    wb.save(response)
    return response


from django.db.models import Prefetch
from django.contrib import messages
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
import datetime

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Prefetch
import datetime

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Prefetch
import datetime

@login_required(login_url='connexion')
def rapport_approvisionnement(request):
    mois = request.GET.get("mois")
    annee = request.GET.get("annee")

    approvisionnements = (
        Approvisionnement.objects
        .prefetch_related(
            Prefetch("appromateriel_set", queryset=ApproMateriel.objects.select_related("materiel"))
        )
        .order_by("-date_appro")
    )

    attributions = (
        Attribution.objects
        .prefetch_related(
            Prefetch("attribumateriel_set", queryset=AttribuMateriel.objects.select_related("materiel"))
        )
        .order_by("-date_attri")
    )

    if mois and annee:
        try:
            mois = int(mois)
            annee = int(annee)
            approvisionnements = approvisionnements.filter(date_appro__month=mois, date_appro__year=annee)
            attributions = attributions.filter(date_attri__month=mois, date_attri__year=annee)
        except ValueError:
            messages.warning(request, "Paramètres de filtre invalides")

    if request.GET.get("export") == "excel":
        return export_excel(approvisionnements, attributions)

    mois_noms = [
        (1, "Janvier"), (2, "Février"), (3, "Mars"), (4, "Avril"),
        (5, "Mai"), (6, "Juin"), (7, "Juillet"), (8, "Août"),
        (9, "Septembre"), (10, "Octobre"), (11, "Novembre"), (12, "Décembre")
    ]

    return render(request, "appro/rapport_approvisionnement.html", {
        "approvisionnements": approvisionnements,
        "attributions": attributions,
        "mois": mois,
        "annee": annee,
        "mois_noms": mois_noms,
        "now": datetime.datetime.now(),
    })



from django.http import HttpResponse
import openpyxl
#=================SORTIE MAGASIN=============================
def export_excel(approvisionnements, attributions):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approvisionnements"

    # Entêtes
    ws.append(["Date", "Technicien", "Matériel", "Quantité"])

    # Données
    for appro in approvisionnements:
        for ligne in appro.appromateriel_set.all():
            # ⚡ Convertir l'objet Materiel en chaîne
            ws.append([
                appro.date_appro.strftime("%Y-%m-%d"),
                appro.utilisateur.nom,
                str(ligne.materiel),  # <-- important !
                ligne.quantite
            ])

    # Préparer la réponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=rapport.xlsx'
    wb.save(response)
    return response

# === EXPORT EXCEL ===
@login_required(login_url='connexion')
def export_appro_excel(appros, attributions):
    if not request.session.get('is_authenticated'):
        messages.warning(request, "Vous devez être connecté pour accéder à cette page.")
        return redirect('/connexion/')
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Approvisionnements"

    # Feuille Approvisionnements
    ws1.append(["Date Appro", "Utilisateur", "Matériel", "Quantité"])
    for appro in appros:
        for ligne in appro.appromateriel_set.all():
            ws1.append([
                appro.date_appro.strftime("%d/%m/%Y %H:%M"),
                appro.utilisateur.nom if appro.utilisateur else "-",
                f"{ligne.materiel.categorie.label} - {ligne.materiel.fabricant} - {ligne.materiel.description} ({ligne.materiel.modal})",
                ligne.quantite,
            ])

    # Feuille Attributions
    ws2 = wb.create_sheet(title="Attributions liées")
    ws2.append(["Date Attribution", "Direction", "Département", "Service", "Bénéficiaire", "Matériel", "Quantité", "État"])
    for attri in attributions:
        for ligne in attri.attribumateriel_set.all():
            ws2.append([
                attri.date_attri.strftime("%d/%m/%Y %H:%M"),
                attri.direction,
                attri.departement,
                attri.service,
                f"{attri.prenom} {attri.nom}",
                f"{ligne.materiel.categorie.label} - {ligne.materiel.fabricant} - {ligne.materiel.description} ({ligne.materiel.modal})",
                ligne.quantite,
                attri.etat,
            ])

    # Ajuster largeur colonnes
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content=openpyxl.writer.excel.save_virtual_workbook(wb),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="rapport_approvisionnement.xlsx"'
    return response
from django.http import JsonResponse
from .models import ApproMateriel

def get_materiels_by_appro(request, appro_id):
    materiels = ApproMateriel.objects.filter(appro_id=appro_id).select_related("materiel")
    data = [
        {"id": m.materiel.id, "text": str(m.materiel)}
        for m in materiels
    ]
    return JsonResponse(data, safe=False)

from django.http import JsonResponse
from .models import Departement, Service

def load_departements(request):
    direction_id = request.GET.get('direction_id')
    departements = Departement.objects.filter(direction_id=direction_id).order_by('nom')
    data = [{"id": d.id, "nom": d.nom} for d in departements]
    return JsonResponse(data, safe=False)

def load_services(request):
    departement_id = request.GET.get('departement_id')
    services = Service.objects.filter(departement_id=departement_id).order_by('nom')
    data = [{"id": s.id, "nom": s.nom} for s in services]
    return JsonResponse(data, safe=False)


from django.shortcuts import render

def handler404(request, exception):
    return render(request, "errors/404.html", status=404)

def handler500(request):
    return render(request, "errors/500.html", status=500)

def handler403(request, exception):
    return render(request, "errors/403.html", status=403)

def handler400(request, exception):
    return render(request, "errors/400.html", status=400)
